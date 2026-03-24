import pandas as pd
import re
from openpyxl import load_workbook

# ---------------------------------
# 1. Scrap de la tabla
# ---------------------------------
url = "https://lto.epri.com/AMPs#International_Aging_Management_Programs"

tables = pd.read_html(url)

df = tables[3].iloc[:, :3]
df.columns = ["Web_Col1", "Web_Col2", "Web_Col3"]

# eliminar duplicados por celdas combinadas
df = df.drop_duplicates(subset=["Web_Col1", "Web_Col2", "Web_Col3"])

# ---------------------------------
# 2. Leer Excel de referencia
# ---------------------------------
excel_path = r"C:\Users\celin\Downloads\IGALL\resultado_WOPI_modificado.xlsx"

df_ref = pd.read_excel(excel_path)
df_ref.iloc[:,0] = df_ref.iloc[:,0].astype(str)

# ---------------------------------
# 3. Buscar AMP###
# ---------------------------------
resultados = []

for _, row in df.iterrows():

    texto = str(row["Web_Col3"])

    amps = re.findall(r"AMP\d+", texto)

    for amp in amps:

        matches = df_ref[df_ref.iloc[:,0].str.startswith(amp, na=False)]

        for _, match in matches.iterrows():

            nueva_fila = {
                "Web_Col1": row["Web_Col1"],
                "Web_Col2": row["Web_Col2"],
                "AMP": amp
            }

            for col in df_ref.columns:
                nueva_fila[col] = match[col]

            resultados.append(nueva_fila)

df_final = pd.DataFrame(resultados)

# ---------------------------------
# 4. Exportar Excel
# ---------------------------------
output = r"C:\Users\celin\Downloads\IGALL\matching_amp_4.xlsx"

df_final.to_excel(output, index=False)

# ---------------------------------
# 5. Combinar celdas columnas 1 y 2
# ---------------------------------
wb = load_workbook(output)
ws = wb.active


def merge_column(col):

    start_row = 2
    merge_start = start_row
    current_value = ws.cell(start_row, col).value

    for row in range(start_row + 1, ws.max_row + 1):

        if ws.cell(row, col).value != current_value:

            if row - merge_start > 1:
                ws.merge_cells(start_row=merge_start,
                               start_column=col,
                               end_row=row-1,
                               end_column=col)

            merge_start = row
            current_value = ws.cell(row, col).value

    if ws.max_row - merge_start >= 1:
        ws.merge_cells(start_row=merge_start,
                       start_column=col,
                       end_row=ws.max_row,
                       end_column=col)

# combinar columnas 1 y 2
merge_column(1)
merge_column(2)
merge_column(3)
merge_column(6)

# convertir URLs en hipervínculos
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

wb.save(output)

print("Archivo generado:", output)