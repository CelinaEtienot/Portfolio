import os
import re
import regex
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font
import json
import urllib.parse

# === CONFIGURACIÓN ===
carpeta = r"C:\Users\celin\Downloads\IGALL"
json_file = r"C:\Users\celin\Downloads\IGALL\ampsjson.txt"

# Patrón flexible de la frase completa (orden obligatorio)
patron_electric = regex.compile(
    r"(ELECTRIC\s+POWER\s+RESEARCH\s+INSTITUTE){e<=1}",  # hasta 1 error en toda la frase
    regex.IGNORECASE
)

# Patrón para EPRI
patron_epri = regex.compile(r"(EPRI){e<=0}", regex.IGNORECASE)

# === LEER JSON ===
with open(json_file, "r", encoding="utf-8") as f:
    data = json.load(f)

file_to_guid = {}
for item in data:
    nombre_archivo = item.get("FileLeafRef")
    guid = item.get("UniqueId", "").strip("{}")
    if nombre_archivo and guid:
        file_to_guid[nombre_archivo] = guid

# === CREAR EXCEL ===
wb = Workbook()
ws = wb.active
ws.title = "Resultados"
ws.append(["Codigo", "Texto encontrado", "Link"])

# === RECORRER ARCHIVOS WORD ===
for archivo in os.listdir(carpeta):
    if archivo.endswith(".docx"):

        ruta_completa = os.path.join(carpeta, archivo)
        nombre_sin_ext = os.path.splitext(archivo)[0]
        codigo = nombre_sin_ext[:6]

        guid = file_to_guid.get(archivo)
        if guid:
            url = f"https://gnssn.iaea.org/NSNI/PoS/IGALL/_layouts/15/WopiFrame.aspx?sourcedoc={urllib.parse.quote('{' + guid + '}')}&action=default"
        else:
            url = ""

        doc = Document(ruta_completa)

        dentro_references = False
        encontrado = False

        for p in doc.paragraphs:
            texto = p.text.strip()
            if not texto:
                continue

            # detectar inicio de REFERENCES
            if re.search(r'\bREFERENCES\b', texto, re.IGNORECASE):
                dentro_references = True
                continue

            if dentro_references:
                # buscar la frase completa en orden o EPRI
                if patron_electric.search(texto) or patron_epri.search(texto) and "preprint" not in texto.lower():
                    ws.append([codigo, texto, url])
                    if url:
                        cell = ws.cell(row=ws.max_row, column=3)
                        cell.hyperlink = url
                        cell.font = Font(color="0000FF", underline="single")
                    encontrado = True

        if not encontrado:
            ws.append([codigo, "", url])
            if url:
                cell = ws.cell(row=ws.max_row, column=3)
                cell.hyperlink = url
                cell.font = Font(color="0000FF", underline="single")

# === GUARDAR EXCEL ===
wb.save(r"C:\Users\celin\Downloads\IGALL\resultado_WOPI.xlsx")
print("Proceso finalizado. Excel creado.")