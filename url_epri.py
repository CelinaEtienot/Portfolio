import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# archivos
archivo_origen = r"C:\Users\celin\Downloads\IGALL\resultado_WOPI_modificado.xlsx"
archivo_nuevo = r"C:\Users\celin\Downloads\IGALL\AMP_references_EPRI.xlsx"

wb = load_workbook(archivo_origen)
ws = wb.active

wb_new = Workbook()
ws_new = wb_new.active

# copiar encabezados
for col, cell in enumerate(ws[1], start=1):
    new_cell = ws_new.cell(row=1, column=col, value=cell.value)

    if cell.hyperlink:
        new_cell.hyperlink = cell.hyperlink.target
        new_cell.font = Font(color="0000FF", underline="single")

# nueva columna URL
url_col = ws.max_column + 1
ws_new.cell(row=1, column=url_col, value="EPRI URL")

patron_inicio = re.compile(
    r'^(?:\[\d+\]\s*)?ELECTRIC\s+POWER\s+RESEARCH\s+INSTITUTE,\s*',
    re.IGNORECASE
)

base_url = "https://www.epri.com/research/products/"

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):

    for col_idx, cell in enumerate(row, start=1):

        valor = cell.value

        # limpiar columna 2
        if col_idx == 2 and isinstance(valor, str):
            valor = patron_inicio.sub("", valor)

        new_cell = ws_new.cell(row=row_idx, column=col_idx, value=valor)

        if cell.hyperlink:
            new_cell.hyperlink = cell.hyperlink.target
            new_cell.font = Font(color="0000FF", underline="single")

    # generar URL desde columna 4
    codigo = row[3].value

    if codigo:
        codigo_str = str(codigo).strip()

        # eliminar prefijo TR-
        codigo_str = re.sub(r'^TR[-\s]*', '', codigo_str, flags=re.IGNORECASE)

        codigo18 = codigo_str.zfill(18)

        url = base_url + codigo18

        cell_url = ws_new.cell(row=row_idx, column=url_col, value=url)
        cell_url.hyperlink = url
        cell_url.font = Font(color="0000FF", underline="single")

# ---------- combinar celdas columna 1 ----------
start = 2
for i in range(3, ws_new.max_row + 2):
    if i > ws_new.max_row or ws_new.cell(i,1).value != ws_new.cell(start,1).value:
        if i - start > 1:
            ws_new.merge_cells(start_row=start, start_column=1, end_row=i-1, end_column=1)
        start = i

# ---------- combinar celdas columna 3 ----------
start = 2
for i in range(3, ws_new.max_row + 2):
    if i > ws_new.max_row or ws_new.cell(i,3).value != ws_new.cell(start,3).value:
        if i - start > 1:
            ws_new.merge_cells(start_row=start, start_column=3, end_row=i-1, end_column=3)
        start = i

wb_new.save(archivo_nuevo)

print("Excel generado:", archivo_nuevo)